import os
from pathlib import Path
from skimage.metrics import structural_similarity as ssim
from PIL import Image
import cv2
import json
import openpyxl

def pathToProject():
    projectPath = os.getcwd()
    return projectPath

def createNewFolder(folder_path):
    currentProjectPath = pathToProject()
    print(currentProjectPath + folder_path)
    Path(currentProjectPath + folder_path).mkdir(exist_ok=True)
    
def compare_images(image1_path, image2_path, threshold=0.9):
    """
    Compare two images using SSIM and return whether they are similar.

    Args:
        image1_path (str): Path to the first image.
        image2_path (str): Path to the second image.
        threshold (float): Similarity threshold (0 to 1). Default is 0.9.

    Returns:
        bool: True if images are similar, False otherwise.
    """
    img1 = cv2.imread(image1_path, cv2.IMREAD_GRAYSCALE)
    img2 = cv2.imread(image2_path, cv2.IMREAD_GRAYSCALE)
    if img1 is None or img2 is None:
        raise FileNotFoundError("One of the images was not found.")
    score, _ = ssim(img1, img2, full=True)
    return score <= threshold

def crop_image(input_image, x1, y1, x2, y2, output_image):
    """
    Crop a specific region from the input image and save it to a new file.

    Args:
        input_image (str): Path to the input image.
        x (int): X-coordinate of the top-left corner of the crop area.
        y (int): Y-coordinate of the top-left corner of the crop area.
        width (int): Width of the crop area.
        height (int): Height of the crop area.
        output_image (str): Path to save the cropped image.
    """
    if not os.path.exists(input_image):
        raise FileNotFoundError(f"Input image {input_image} does not exist.")
    try:
        image = Image.open(input_image)
    except Exception as e:
        raise ValueError(f"Error opening image {input_image}: {e}")
    width= x2-x1
    height= y2-y1
    if x1 < 0 or y1 < 0 or x1 + width > image.width or y1 + height > image.height:
        raise ValueError(f"Invalid crop dimensions: ({x1}, {y1}, {width}, {height})")
    cropped = image.crop((x1, y1, x1 + width, y1 + height))
    try:
        cropped.save(output_image)
        print(f"Cropped image saved as {output_image}")
    except Exception as e:
        raise ValueError(f"Error saving cropped image: {e}")
def get_one_element(Sheetname, row, cell):
    row=int(row)
    cell=int(row)
    currentDir = os.getcwd()
    file_path = os.path.join(currentDir, "utils", "test_xcel.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    testData = openpyxl.load_workbook(file_path)
    if Sheetname not in testData.sheetnames:
        raise ValueError(f"Sheet '{Sheetname}' does not exist in the Excel file")
    sh = testData[Sheetname]
    cell_value = sh.cell(row, cell).value
    if cell_value is None:
        raise ValueError(f"The cell at row {row}, column {cell} is empty")
    return cell_value

def convert_xlsx_to_json(sheet_name,index,attribute):
    # Đọc file Excel
    currentDir = os.getcwd()
    index = int(index)
    file_path = os.path.join(currentDir, "utils", "multi_language.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the Excel file")
    sheet = workbook[sheet_name]
    #Get header
    headers = [cell.value for cell in sheet[1][:5]] 
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True): 
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        if row_data.get("ID") is not None: 
            data.append(row_data)
    if data:
        return json.dumps(data[index].get(f'{attribute}'), ensure_ascii=False, indent=4)
    else:
        return None
def get_all_value_with_attribute(sheet_name, attribute):
    # Đọc file Excel
    currentDir = os.getcwd()
    file_path = os.path.join(currentDir, "utils", "multi_language.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the Excel file")
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    if f"{attribute}" not in headers:
        raise ValueError(f"Column '{attribute}' does not exist in the sheet")
    attribute_col_index = headers.index(f"{attribute}")
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True): 
        if row[attribute_col_index] is not None:  # Lọc các hàng có giá trị ở cột 'Vietnamese'
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            data.append(row_data)
    if data:
        return json.dumps([item.get(attribute) for item in data], ensure_ascii=False, indent=4)
    else:
        return f"No data found for the column '{attribute}'."
    
def convert_excel_to_custom_json(sheet_name):
    currentDir = os.getcwd()
    file_path = os.path.join(currentDir, "utils", "language.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    sheet = workbook[sheet_name]
    result = {sheet_name: {}}
    for row in sheet.iter_rows(min_row=2, values_only=True): 
        screen = row[1]  
        name_id = row[2]  
        language = row[3] 
        if screen is None:
            break
        # Tạo cấu trúc JSON
        if screen not in result[sheet_name]:
            result[sheet_name][screen] = {}
        result[sheet_name][screen][name_id] = language
    output_file_path = os.path.join(currentDir, "utils", f"{sheet_name}.json")
    # Xuất ra file JSON
    with open(output_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(result, json_file, ensure_ascii=False, indent=4)
    # try:
    #     specific_value = result[sheet_name][f"{screen_name}"][f"{id}"]
    #     return specific_value  # Trả về giá trị cụ thể
    # except KeyError as e:
    #     raise ValueError(f"Không tìm thấy key cụ thể: {e}")
